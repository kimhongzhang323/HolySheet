from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from typing import Optional
from bson import ObjectId
from datetime import datetime, timedelta
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from ..db import get_database
from ..models.user import UserResponse, UserRole
from ..dependencies import get_current_user

router = APIRouter()

@router.get("/admin/reports/weekly")
async def generate_weekly_report(
    start_date: Optional[str] = None,
    current_user: UserResponse = Depends(get_current_user),
    db = Depends(get_database)
):
    """Generate weekly summary report"""
    if current_user.role not in [UserRole.ADMIN, UserRole.STAFF]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only admin or staff can generate reports"
        )
    
    # Default to current week if no start_date provided
    if start_date:
        try:
            week_start = datetime.strptime(start_date, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    else:
        # Get Monday of current week
        today = datetime.utcnow()
        week_start = today - timedelta(days=today.weekday())
    
    week_end = week_start + timedelta(days=7)
    
    # Get activities for the week
    query = {
        "start_time": {"$gte": week_start, "$lt": week_end}
    }
    
    cursor = db.activities.find(query).sort("start_time", 1)
    activities = await cursor.to_list(length=500)
    
    # Calculate stats
    total_activities = len(activities)
    total_capacity = sum(act.get("capacity", 0) for act in activities)
    total_attended = sum(len(act.get("attendees", [])) for act in activities)
    
    # Volunteer stats
    total_volunteers_needed = sum(act.get("volunteers_needed", 0) for act in activities)
    total_volunteers_registered = sum(act.get("volunteers_registered", 0) for act in activities)
    
    # Average attendance rate
    avg_attendance_rate = (total_attended / total_capacity * 100) if total_capacity > 0 else 0
    
    # Volunteer fulfillment rate
    volunteer_fulfillment_rate = (total_volunteers_registered / total_volunteers_needed * 100) if total_volunteers_needed > 0 else 0
    
    return {
        "week_start": week_start.strftime("%Y-%m-%d"),
        "week_end": week_end.strftime("%Y-%m-%d"),
        "total_activities": total_activities,
        "total_capacity": total_capacity,
        "total_attended": total_attended,
        "avg_attendance_rate": round(avg_attendance_rate, 1),
        "total_volunteers_needed": total_volunteers_needed,
        "total_volunteers_registered": total_volunteers_registered,
        "volunteer_fulfillment_rate": round(volunteer_fulfillment_rate, 1),
        "activities": [
            {
                "title": act["title"],
                "date": act["start_time"].strftime("%Y-%m-%d %H:%M"),
                "location": act.get("location"),
                "capacity": act.get("capacity", 0),
                "attended": len(act.get("attendees", [])),
                "volunteers_needed": act.get("volunteers_needed", 0),
                "volunteers_registered": act.get("volunteers_registered", 0)
            }
            for act in activities
        ]
    }


@router.get("/admin/reports/activity/{activity_id}/export")
async def export_activity_csv(
    activity_id: str,
    current_user: UserResponse = Depends(get_current_user),
    db = Depends(get_database)
):
    """Export activity details as CSV"""
    if current_user.role not in [UserRole.ADMIN, UserRole.STAFF]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only admin or staff can export reports"
        )
    
    if not ObjectId.is_valid(activity_id):
        raise HTTPException(status_code=400, detail="Invalid activity ID")
    
    activity = await db.activities.find_one({"_id": ObjectId(activity_id)})
    if not activity:
        raise HTTPException(status_code=404, detail="Activity not found")
    
    # Get attendee details
    attendee_ids = activity.get("attendees", [])
    attendees = []
    
    for attendee_id in attendee_ids:
        if ObjectId.is_valid(attendee_id):
            user = await db.users.find_one({"_id": ObjectId(attendee_id)})
            if user:
                attendees.append({
                    "name": user.get("name"),
                    "email": user.get("email"),
                    "phone": user.get("phoneNumber", "N/A")
                })
    
    # Generate CSV
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow([f"Activity Report: {activity['title']}"])
    writer.writerow([f"Date: {activity['start_time'].strftime('%Y-%m-%d %H:%M')}"])
    writer.writerow([f"Location: {activity.get('location', 'N/A')}"])
    writer.writerow([])
    writer.writerow(["Name", "Email", "Phone"])
    
    # Attendees
    for attendee in attendees:
        writer.writerow([attendee["name"], attendee["email"], attendee["phone"]])
    
    # Return as downloadable file
    output.seek(0)
    headers = {
        'Content-Disposition': f'attachment; filename="activity_{activity_id}.csv"'
    }
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers=headers
    )


@router.get("/admin/reports/volunteers/export")
async def export_volunteers_excel(
    current_user: UserResponse = Depends(get_current_user),
    db = Depends(get_database)
):
    """Export volunteer roster as Excel"""
    if current_user.role not in [UserRole.ADMIN, UserRole.STAFF]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only admin or staff can export reports"
        )
    
    # Get all volunteers
    cursor = db.users.find({"role": {"$in": [UserRole.VOLUNTEER, "volunteer"]}})
    volunteers = await cursor.to_list(length=1000)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Volunteers"
    
    # Header styling
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Headers
    headers = ["Name", "Email", "Phone", "Skills", "Tier"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Data
    for row, vol in enumerate(volunteers, 2):
        ws.cell(row=row, column=1, value=vol.get("name", ""))
        ws.cell(row=row, column=2, value=vol.get("email", ""))
        ws.cell(row=row, column=3, value=vol.get("phoneNumber", ""))
        ws.cell(row=row, column=4, value=", ".join(vol.get("skills", [])))
        ws.cell(row=row, column=5, value=vol.get("tier", ""))
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    headers = {
        'Content-Disposition': 'attachment; filename="volunteers_roster.xlsx"'
    }
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )
